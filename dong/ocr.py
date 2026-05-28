"""dong 的本地 OCR 适配层，负责把图片转成可发送给文本模型的纯文本。"""

from __future__ import annotations

import os
import re
import zipfile
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

ATTACHMENT_MARKER_RE = re.compile(
    r"\[(?:image|attachment|file):(?P<path>[^\]\r\n]+)\]"
)
IMAGE_MARKER_RE = ATTACHMENT_MARKER_RE
_MAX_EXTRACTED_TEXT_CHARS = 40_000


class OcrError(RuntimeError):
    """OCR 识别失败时抛出的用户可读异常。"""


@dataclass(frozen=True)
class OcrLine:
    """单行 OCR 结果，保留文本、置信度和原始框信息。"""

    text: str
    score: float | None = None
    box: Any | None = None


@dataclass(frozen=True)
class OcrResult:
    """一次图片 OCR 的结构化结果。"""

    image_path: str
    lines: tuple[OcrLine, ...]

    @property
    def text(self) -> str:
        """按识别顺序拼接文本，供后续 LLM prompt 使用。"""
        return "\n".join(line.text for line in self.lines if line.text.strip())


def build_ocr_prompt(result: OcrResult, question: str = "") -> str:
    """把 OCR 文本包装成普通文本 prompt，避免依赖模型原生图片能力。"""
    return build_ocr_prompt_for_results((result,), question)


def build_ocr_prompt_for_results(
    results: tuple[OcrResult, ...],
    question: str = "",
) -> str:
    """把一份或多份附件的提取文本包装成普通文本 prompt。"""
    user_question = question.strip()
    attachment_sections = []
    for index, result in enumerate(results, start=1):
        ocr_text = result.text.strip() or "(OCR 未识别到文字)"
        attachment_sections.append(
            "\n".join([
                f"附件 {index} 提取文本：",
                ocr_text,
            ])
        )

    parts = [
        "用户上传了附件。下面是本地提取出的附件文字，供你理解内容；不要假装直接看到了原始文件。",
        *attachment_sections,
    ]
    if user_question:
        parts.extend(["用户输入：", user_question])
    return "\n\n".join(parts)


def image_marker(path: str | Path) -> str:
    """生成输入框中可见的图片占位符。"""
    return f"[image:{Path(path).expanduser()}]"


def attachment_marker(path: str | Path) -> str:
    """生成输入框中可见的通用附件占位符。"""
    return f"[attachment:{Path(path).expanduser()}]"


def expand_image_markers_to_ocr_prompt(text: str) -> str | None:
    """把输入中的附件占位符转成文本 prompt；没有附件时返回 None。"""
    paths = tuple(
        match.group("path").strip() for match in ATTACHMENT_MARKER_RE.finditer(text)
    )
    if not paths:
        return None

    question = ATTACHMENT_MARKER_RE.sub(" ", text).strip()
    results = tuple(extract_text_from_attachment(path) for path in paths)
    return build_ocr_prompt_for_results(results, question)


def expand_attachment_markers_to_prompt(text: str) -> str | None:
    """把输入中的附件占位符转成文本 prompt；没有附件时返回 None。"""
    return expand_image_markers_to_ocr_prompt(text)


def extract_text_from_attachment(path: str) -> OcrResult:
    """从图片、PDF、Excel、Word 或 PowerPoint 附件中提取可读文字。"""
    file_path = Path(path).expanduser().resolve()
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return _extract_text_from_pdf(file_path)
    if suffix == ".xlsx":
        return _extract_text_from_xlsx(file_path)
    if suffix == ".docx":
        return _extract_text_from_docx(file_path)
    if suffix == ".pptx":
        return _extract_text_from_pptx(file_path)
    return _do_extract(path)


def extract_text_from_image(path: str) -> OcrResult:
    """使用 OnnxOCR 在 CPU 上识别图片文字。"""
    return _do_extract(path)


def _do_extract(path: str) -> OcrResult:
    """对单张图片执行本地 OCR。"""
    image_path = Path(path).expanduser().resolve()
    if not image_path.is_file():
        raise OcrError(f"图片文件不存在：{image_path}")

    cv2 = _load_cv2()
    image = cv2.imread(str(image_path))
    if image is None:
        raise OcrError(f"无法读取图片文件：{image_path}")

    raw_result = _onnxocr_model().ocr(image, cls=False)
    lines = tuple(_iter_ocr_lines(raw_result))
    return OcrResult(image_path=str(image_path), lines=lines)


def _extract_text_from_pdf(pdf_path: Path) -> OcrResult:
    """把 PDF 转成页面图片后逐页 OCR。"""
    if not pdf_path.is_file():
        raise OcrError(f"PDF 文件不存在：{pdf_path}")
    images = _pdf_to_images(pdf_path)
    if not images:
        raise OcrError(f"PDF 未解析出页面：{pdf_path}")

    lines: list[OcrLine] = []
    for page_index, image in enumerate(images, start=1):
        lines.append(OcrLine(text=f"第 {page_index} 页："))
        raw_result = _onnxocr_model().ocr(image, cls=False)
        page_lines = tuple(_iter_ocr_lines(raw_result))
        if page_lines:
            lines.extend(page_lines)
        else:
            lines.append(OcrLine(text="(OCR 未识别到文字)"))
    return OcrResult(image_path=str(pdf_path), lines=tuple(lines))


def _pdf_to_images(pdf_path: Path) -> list[Any]:
    """调用 OnnxOCR 自带 PDF 转图实现。"""
    try:
        from onnxocr.ocr_images_pdfs import pdf_to_images  # type: ignore[import-not-found]
    except ImportError as exc:
        raise OcrError("缺少 PDF OCR 运行依赖；请先执行：uv sync") from exc
    if pdf_to_images is None:
        raise OcrError("缺少 PDF 转图片依赖 pymupdf；请先执行：uv sync")
    return list(pdf_to_images(str(pdf_path), dpi=300))


def _extract_text_from_xlsx(file_path: Path) -> OcrResult:
    """读取 Excel 工作簿中的单元格文本。"""
    if not file_path.is_file():
        raise OcrError(f"Excel 文件不存在：{file_path}")
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise OcrError("缺少 Excel 读取依赖 openpyxl；请先执行：uv sync") from exc

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise OcrError(f"无法读取 Excel 文件：{file_path.name}") from exc
    try:
        sections: list[str] = []
        for sheet in workbook.worksheets:
            rows: list[str] = []
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None]
                if values:
                    rows.append("\t".join(values))
            if rows:
                sections.append("\n".join([f"工作表：{sheet.title}", *rows]))
        text = "\n\n".join(sections)
    finally:
        workbook.close()
    return _text_result(file_path, text or "(未提取到 Excel 文本)")


def _extract_text_from_docx(file_path: Path) -> OcrResult:
    """从 Word OOXML 文档中提取正文文本。"""
    text = _extract_zip_xml_text(
        file_path,
        path_predicate=lambda name: name == "word/document.xml",
    )
    return _text_result(file_path, text or "(未提取到 Word 文本)")


def _extract_text_from_pptx(file_path: Path) -> OcrResult:
    """从 PowerPoint OOXML 幻灯片中提取文本。"""
    text = _extract_zip_xml_text(
        file_path,
        path_predicate=lambda name: name.startswith("ppt/slides/slide")
        and name.endswith(".xml"),
    )
    return _text_result(file_path, text or "(未提取到 PowerPoint 文本)")


def _extract_zip_xml_text(file_path: Path, *, path_predicate) -> str:
    """读取 Office zip 包内指定 XML 文件的文本节点。"""
    if not file_path.is_file():
        raise OcrError(f"文件不存在：{file_path}")
    try:
        with zipfile.ZipFile(file_path) as archive:
            text_parts: list[str] = []
            for name in sorted(archive.namelist()):
                if not path_predicate(name):
                    continue
                text_parts.extend(_iter_xml_text(archive.read(name)))
    except (OSError, zipfile.BadZipFile, ElementTree.ParseError) as exc:
        raise OcrError(f"无法读取 Office 文件：{file_path.name}") from exc
    return "\n".join(part for part in text_parts if part.strip())


def _iter_xml_text(xml_bytes: bytes):
    """遍历 OOXML 中的文本节点。"""
    root = ElementTree.fromstring(xml_bytes)
    for element in root.iter():
        if element.tag.rsplit("}", 1)[-1] == "t" and element.text:
            yield element.text.strip()


def _text_result(file_path: Path, text: str) -> OcrResult:
    """把普通文本包装成 OcrResult，复用后续 prompt 结构。"""
    normalized = _truncate_text(text.strip())
    lines = tuple(OcrLine(line) for line in normalized.splitlines() if line.strip())
    return OcrResult(image_path=str(file_path), lines=lines or (OcrLine("(未提取到文本)"),))


def _truncate_text(text: str) -> str:
    """限制文档提取文本长度，避免一次附件塞爆上下文。"""
    if len(text) <= _MAX_EXTRACTED_TEXT_CHARS:
        return text
    return text[:_MAX_EXTRACTED_TEXT_CHARS] + "\n...(文本过长，已截断)"


def _load_cv2():
    """延迟导入 OpenCV，让未使用 OCR 的启动路径不受可选依赖影响。"""
    try:
        import cv2  # type: ignore[import-not-found]
    except ImportError as exc:
        raise OcrError(
            "缺少 OCR 运行依赖 opencv-python-headless；"
            "请先执行：uv sync"
        ) from exc
    return cv2


@lru_cache(maxsize=1)
def _onnxocr_model():
    """懒加载 OnnxOCR 模型，避免每条命令重复初始化 ONNXRuntime。"""
    try:
        from onnxocr.onnx_paddleocr import ONNXPaddleOcr  # type: ignore[import-not-found]
    except ImportError as exc:
        raise OcrError("缺少 OnnxOCR；请先执行：uv sync") from exc

    use_angle_cls = os.getenv("DONG_OCR_USE_ANGLE_CLS", "").lower() in {
        "1",
        "true",
        "yes",
    }
    return ONNXPaddleOcr(use_angle_cls=use_angle_cls, use_gpu=False)


def _iter_ocr_lines(raw_result: Any):
    """兼容 OnnxOCR/PaddleOCR 常见嵌套结构并提取文本行。"""
    if not raw_result:
        return

    pages = raw_result if isinstance(raw_result, list) else [raw_result]
    for page in pages:
        if not page:
            continue
        for item in page:
            line = _parse_line(item)
            if line is not None and line.text.strip():
                yield line


def _parse_line(item: Any) -> OcrLine | None:
    """解析单条 OCR 结果；无法识别的结构直接跳过。"""
    if not isinstance(item, (list, tuple)) or len(item) < 2:
        return None

    box = item[0]
    rec = item[1]
    if isinstance(rec, (list, tuple)) and rec:
        text = str(rec[0] or "")
        score = _to_float(rec[1]) if len(rec) > 1 else None
        return OcrLine(text=text, score=score, box=box)
    if isinstance(rec, str):
        return OcrLine(text=rec, box=box)
    return None


def _to_float(value: Any) -> float | None:
    """把 numpy 标量等置信度值安全转换为 float。"""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
